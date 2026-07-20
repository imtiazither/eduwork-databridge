import csv
import hashlib
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast

import pyarrow.parquet as pq
from openpyxl import load_workbook

from eduwork_databridge.connectors.base import (
    ConnectionTestResult,
    Connector,
    ConnectorError,
    DiscoveryResult,
    ExtractionBatch,
    FieldDiscovery,
)
from eduwork_databridge.connectors.security import ensure_allowed_file, validate_zip_archive
from eduwork_databridge.schemas.config import SourceObjectConfig


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _observed_type(values: list[Any]) -> str:
    non_null = [value for value in values if value not in (None, "")]
    if not non_null:
        return "unknown"
    types = {type(value).__name__ for value in non_null}
    return next(iter(types)) if len(types) == 1 else "mixed"


def discover_fields(records: list[dict[str, Any]], object_key: str) -> DiscoveryResult:
    sample = records[:100]
    names: list[str] = []
    for record in sample:
        for name in record:
            if name not in names:
                names.append(name)
    fields = [
        FieldDiscovery(
            name=name,
            observed_type=_observed_type([record.get(name) for record in sample]),
            nullable=any(record.get(name) in (None, "") for record in sample),
        )
        for name in names
    ]
    return DiscoveryResult(object_key=object_key, fields=fields, sample_count=len(sample))


class FileConnector(Connector):
    suffixes: set[str]
    content_type: str

    def __init__(self, allowed_roots: list[Path], max_bytes: int) -> None:
        self.allowed_roots = allowed_roots
        self.max_bytes = max_bytes

    async def test_connection(self) -> ConnectionTestResult:
        existing = sum(root.expanduser().exists() for root in self.allowed_roots)
        return ConnectionTestResult(
            ok=existing == len(self.allowed_roots),
            connector=self.connector_type,
            checked_at=datetime.now(UTC),
            details={"configured_roots": len(self.allowed_roots), "existing_roots": existing},
        )

    def _path(self, source_object: SourceObjectConfig) -> Path:
        path = ensure_allowed_file(
            Path(source_object.location), self.allowed_roots, self.max_bytes, self.suffixes
        )
        if path.suffix.lower() == ".xlsx":
            validate_zip_archive(path)
        return path

    def _read_records(self, path: Path, source_object: SourceObjectConfig) -> list[dict[str, Any]]:
        raise NotImplementedError

    async def extract(
        self,
        source_object: SourceObjectConfig,
        cursor: dict[str, Any] | None = None,
    ) -> ExtractionBatch:
        path = self._path(source_object)
        raw = path.read_bytes()
        checksum = _sha256_bytes(raw)
        current_cursor = {"sha256": checksum, "mtime_ns": path.stat().st_mtime_ns}
        unchanged = cursor is not None and cursor.get("sha256") == checksum
        records = [] if unchanged else self._read_records(path, source_object)
        return ExtractionBatch(
            records=records,
            raw_bytes=raw,
            content_type=self.content_type,
            cursor=current_cursor,
            metadata={
                "unchanged": unchanged,
                "source_bytes": len(raw),
                "row_count": len(records),
            },
        )

    async def discover_schema(self, source_object: SourceObjectConfig) -> DiscoveryResult:
        path = self._path(source_object)
        records = self._read_records(path, source_object)
        return discover_fields(records, source_object.key)


class CSVConnector(FileConnector):
    connector_type = "csv"
    suffixes = {".csv"}
    content_type = "text/csv"

    def _read_records(self, path: Path, source_object: SourceObjectConfig) -> list[dict[str, Any]]:
        delimiter = str(source_object.options.get("delimiter", ","))
        encoding = str(source_object.options.get("encoding", "utf-8"))
        with path.open("r", encoding=encoding, newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle, delimiter=delimiter)]


class JSONConnector(FileConnector):
    connector_type = "json"
    suffixes = {".json"}
    content_type = "application/json"

    def _read_records(self, path: Path, source_object: SourceObjectConfig) -> list[dict[str, Any]]:
        value: Any = json.loads(path.read_text(encoding="utf-8"))
        if source_object.json_records_path:
            for part in source_object.json_records_path.split("."):
                if not isinstance(value, dict) or part not in value:
                    raise ConnectorError(
                        "json_records_path_missing", "JSON records path was not found"
                    )
                value = value[part]
        if not isinstance(value, list) or not all(isinstance(item, dict) for item in value):
            raise ConnectorError(
                "json_records_required", "JSON source must resolve to a list of objects"
            )
        return [dict(item) for item in value]


class XLSXConnector(FileConnector):
    connector_type = "xlsx"
    suffixes = {".xlsx"}
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _read_records(self, path: Path, source_object: SourceObjectConfig) -> list[dict[str, Any]]:
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook[source_object.sheet_name] if source_object.sheet_name else workbook.active
        if sheet is None:
            raise ConnectorError(
                "worksheet_missing", "Workbook does not contain a readable worksheet"
            )
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value) for value in next(iterator)]
        except StopIteration:
            return []
        records = [dict(zip(headers, row, strict=False)) for row in iterator]
        workbook.close()
        return records


class ParquetConnector(FileConnector):
    connector_type = "parquet"
    suffixes = {".parquet"}
    content_type = "application/vnd.apache.parquet"

    def _read_records(self, path: Path, source_object: SourceObjectConfig) -> list[dict[str, Any]]:
        del source_object
        return cast(list[dict[str, Any]], pq.read_table(path).to_pylist())
