import csv
import json
from pathlib import Path
from typing import Any, cast
from urllib.parse import unquote, urlparse

import pyarrow.parquet as pq
from openpyxl import load_workbook

from eduwork_databridge.connectors.base import ConnectorError
from eduwork_databridge.db.models.control import RawSnapshot


def _file_path(storage_uri: str) -> Path:
    parsed = urlparse(storage_uri)
    if parsed.scheme != "file":
        raise ConnectorError("unsupported_snapshot_uri", "Snapshot URI is not a local file")
    return Path(unquote(parsed.path)).resolve(strict=True)


def read_snapshot_records(snapshot: RawSnapshot) -> list[dict[str, Any]]:
    path = _file_path(snapshot.storage_uri)
    content_type = str(snapshot.manifest_json.get("content_type", ""))
    if content_type == "text/csv":
        with path.open(encoding="utf-8", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if content_type in {"application/json", "application/x-ndjson"}:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, list) or not all(isinstance(item, dict) for item in payload):
            raise ConnectorError("snapshot_records_invalid", "Snapshot must contain object records")
        return [dict(item) for item in payload]
    if content_type == "application/vnd.apache.parquet":
        return cast(list[dict[str, Any]], pq.read_table(path).to_pylist())
    if content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet = workbook.active
        if sheet is None:
            raise ConnectorError("worksheet_missing", "Snapshot workbook has no worksheet")
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value) for value in next(rows)]
        except StopIteration:
            return []
        records = [dict(zip(headers, row, strict=False)) for row in rows]
        workbook.close()
        return records
    raise ConnectorError(
        "snapshot_content_type_unsupported", "Snapshot content type is unsupported"
    )
